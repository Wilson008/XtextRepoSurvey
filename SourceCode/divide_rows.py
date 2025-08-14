from openpyxl import Workbook

# 创建新的Excel工作簿
wb = Workbook()
ws = wb.active

# 读取原文件
with open("SourceCode/repos_real_created_time_20250728_070052.csv", 'r', encoding='utf-8') as file:
    row_num = 1
    
    for line in file:
        line = line.strip()  # 去除换行符
        if line:  # 跳过空行
            # 按逗号分割字符串
            parts = line.split(',')
            
            if len(parts) == 4:
                # 将4个子字符串分别写入A、B、C、D列
                ws.cell(row=row_num, column=1, value=parts[0])  # A列
                ws.cell(row=row_num, column=2, value=parts[1])  # B列
                ws.cell(row=row_num, column=3, value=parts[2])  # C列
                ws.cell(row=row_num, column=4, value=parts[3])  # D列
                
                row_num += 1

# 保存为xlsx文件
wb.save("SourceCode/repos_created_time_new.xlsx")

print(f"处理完成！共处理 {row_num-1} 行，已保存为 repos_created_time_20250728.xlsx")