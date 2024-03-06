import requests
import csv
from openpyxl import load_workbook

def get_ecore_files(owner, repo, access_token):
    headers = {
        'Authorization': f'token {access_token}'
    }
    url = f'https://api.github.com/repos/{owner}/{repo}/git/trees/master?recursive=1'
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        response_json = response.json()
        ecore_files = [file.get('path') for file in response_json.get('tree', []) if file.get('type') == 'blob' and file.get('path', '').endswith('.ecore')]
        return ecore_files
    else:
        print(f"Failed to retrieve Ecore files for {owner}/{repo}. Status code: {response.status_code}")
        return None

def write_to_csv(owner, repo, ecore_count):
    with open('count_ecore_in_repos.csv', 'a', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['Owner', 'Repo', 'Ecore File Count']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        if csvfile.tell() == 0:
            writer.writeheader()
        
        writer.writerow({'Owner': owner, 'Repo': repo, 'Ecore File Count': ecore_count})

def main():
    access_token = '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'

    # Load Excel file
    wb = load_workbook('analysis_results/analyze_ecore_xtext_files.xlsx')
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_col=2, max_row=ws.max_row, values_only=True):
        owner = row[0]
        repo = row[1]

        ecore_files = get_ecore_files(owner, repo, access_token)
        if ecore_files:
            ecore_count = len(ecore_files)
            write_to_csv(owner, repo, ecore_count)
            print(f"Found {ecore_count} Ecore files in the repository {owner}/{repo}.")
        else:
            ecore_count = 0
            write_to_csv(owner, repo, ecore_count)
            print(f"No Ecore files found in the repository {owner}/{repo}.")

    print("All repositories queried and results saved.")

if __name__ == "__main__":
    main()
