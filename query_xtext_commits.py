import requests
import csv
from openpyxl import load_workbook

def get_xtext_files(owner, repo, access_token):
    headers = {
        'Authorization': f'token {access_token}'
    }
    url = f'https://api.github.com/repos/{owner}/{repo}/git/trees/master?recursive=1'
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        response_json = response.json()
        xtext_files = []
        for file in response_json.get('tree', []):
            if file.get('type') == 'blob' and file.get('path', '').endswith('.xtext'):
                xtext_files.append(file.get('path'))
        return xtext_files
    else:
        print(f"Failed to retrieve XText files for {owner}/{repo}. Status code: {response.status_code}")
        return None

def get_commit_count(owner, repo, file_path, access_token):
    headers = {
        'Authorization': f'token {access_token}'
    }
    url = f'https://api.github.com/repos/{owner}/{repo}/commits?path={file_path}'
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return len(response.json())
    else:
        print(f"Failed to retrieve commit count for {file_path} in {owner}/{repo}. Status code: {response.status_code}")
        return None

def calculate_average_commit_count(owner, repo, xtext_files, access_token):
    total_commit_count = 0
    failed_requests = 0
    if xtext_files is None:
        return 'null'
    for file_path in xtext_files:
        commit_count = get_commit_count(owner, repo, file_path, access_token)
        if commit_count is None:
            failed_requests += 1
        else:
            total_commit_count += commit_count
    if failed_requests == len(xtext_files):
        return 'failed'
    elif len(xtext_files) - failed_requests == 0:
        return 'null'
    else:
        return round(total_commit_count / (len(xtext_files) - failed_requests), 2)

def main():
    access_token = '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'

    # Load Excel file
    wb = load_workbook('analysis_results/analyze_ecore_xtext_files.xlsx')
    ws = wb.active

    # Write to first CSV file
    with open('xtext_commit_counts.csv', 'w', newline='', encoding='utf-8') as csvfile1:
        fieldnames1 = ['Owner', 'Repo', 'XText File', 'Commit Count']
        writer1 = csv.DictWriter(csvfile1, fieldnames=fieldnames1)
        writer1.writeheader()

        # Write to second CSV file
        with open('xtext_average_commit_count.csv', 'w', newline='', encoding='utf-8') as csvfile2:
            fieldnames2 = ['Owner', 'Repo', 'Average Commit Count']
            writer2 = csv.DictWriter(csvfile2, fieldnames=fieldnames2)
            writer2.writeheader()

            for row in ws.iter_rows(min_row=2, max_col=2, max_row=ws.max_row, values_only=True):
                owner = row[0]
                repo = row[1]

                xtext_files = get_xtext_files(owner, repo, access_token)
                average_commit_count = calculate_average_commit_count(owner, repo, xtext_files, access_token)

                print(f"Querying repo: {owner}/{repo}")

                # Write to first CSV file
                if xtext_files is None:
                    writer1.writerow({'Owner': owner, 'Repo': repo, 'XText File': 'null', 'Commit Count': 'null'})
                else:
                    for file_path in xtext_files:
                        commit_count = get_commit_count(owner, repo, file_path, access_token)
                        if commit_count is None:
                            writer1.writerow({'Owner': owner, 'Repo': repo, 'XText File': file_path, 'Commit Count': 'failed'})
                        else:
                            writer1.writerow({'Owner': owner, 'Repo': repo, 'XText File': file_path, 'Commit Count': commit_count})

                # Write to second CSV file
                writer2.writerow({'Owner': owner, 'Repo': repo, 'Average Commit Count': average_commit_count})

    print("All repositories queried and results saved.")

if __name__ == "__main__":
    main()
